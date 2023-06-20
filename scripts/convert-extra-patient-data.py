import sys
import openpyxl
import sqlite3
from unidecode import unidecode

main_table = "Patients"

fields = {
    "age": {"fields": "edad"},
    "clinic_history": {"fields": "historia clinica"},
    "gender": {"fields": "sexo", "fn": lambda s: {"F": 1, "M": 2}[s]},
    "weight": {"fields": "peso"},
    "height": {"fields": "talla"},
    "bmi": {"fields": "imc"},
    "creatinin": {"fields": "creatinina", "fn": lambda s: s},
    "tsh": {"fields": "tsh/t4 libre", "fn": lambda s: s},
    "max_blood_pressure": {"fields": "tension arterial", "fn": lambda s: int(float(s.split("/")[0]))},
    "min_blood_pressure": {"fields": "tension arterial", "fn": lambda s: int(float(s.split("/")[1]))},
    "diagnostic": {"fields": "diagnostico", "multiline": True, "fn": lambda s: s},
    "base_treatment": {"fields": "tratamiento", "multiline": True, "fn": lambda s: s},
    "extra_studies": {"fields": "estudios", "multiline": True, "fn": lambda s: s},
}


def clean_str(s: str) -> str:
    if s:
        return unidecode(s).lstrip().rstrip().lower()
    
    return ""

def main(filename: str, db: str):
    connection = sqlite3.connect(db)
    cursor = connection.cursor()

    wb = openpyxl.load_workbook(filename=filename, data_only=True)
    sheetnames = [s for s in wb.sheetnames if s.isdigit()]
    
    for s in sheetnames:
        sheet = wb[s]
        sheet_values = tuple(sheet.values)
        patient_id = int(s)
        
        data = {}

        for key, options in fields.items():
            value = ""

            for row in sheet_values:
                if not options["fields"] in clean_str(row[0]):
                    continue
                
                if "multiline" in options and options["multiline"]:
                    i = 1
                    while i < len(row) and row[i]:
                        value += row[i]
                        value += "\n"

                        i += 1

                    value = value.rstrip("\n")
                else:
                    value = row[1]
                
                if value and (isinstance(value, float) or isinstance(value, int) or clean_str(value) not in ["s/d", "n/d", "n/a", "pend", "normal"]):
                    fn = lambda s: int(float(s))

                    if "fn" in options:
                        fn = options["fn"]

                    value = fn(value)
                else:
                    value = ""

                break

            data[key] = value
        
        operations = [f"{key} = '{value}'" for key, value in data.items()]
        operation_string = ", ".join(operations).rstrip(", ")

        cursor.execute(f"UPDATE {main_table} SET {operation_string} WHERE id = {patient_id};")

    connection.commit()


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("ERROR: Incorrect usage")
        print("Usage: python3 convert-extra-patient-data.py [file_name] [database]")
        exit()
    
    filename = sys.argv[1]
    db = sys.argv[2]

    main(filename, db)
