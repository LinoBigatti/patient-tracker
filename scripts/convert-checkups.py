import sys
import datetime
import openpyxl
import csv
from unidecode import unidecode

MAGIC_DATA_STRING = "evaluacion del sintoma a tratar segun escala correspondiente"
MAGIC_DATE_STRING = "fecha"
MAGIC_DOSG_STRING = "dosis"
MAGIC_FREQ_STRING = "frecuencia"
MAGIC_EFFS_STRING = "eventos adversos"


def parse_ranges(v):
    if isinstance(v, datetime.datetime):
        # x-y is parsed by excel as day x, month y
        return v.month
    else:
        return v

def clean_str(s: str) -> str:
    if s:
        return unidecode(s).lstrip().rstrip().lower()
    
    return ""

def main(filename: str, trim_length: int = -1):
    wb = openpyxl.load_workbook(filename=filename, data_only=True)
    sheetnames = [s for s in wb.sheetnames if s.isdigit()]
    
    data = {}
    keys = set()
    for s in sheetnames:
        sheet = wb[s]
        sheet_values = tuple(sheet.values)
        patient_id = int(s)
        date = ""
        freq = ""
        dosg = ""
        effs = ""
        obvs = ""
        
        data[patient_id] = []
        for (i, row) in enumerate(sheet_values):
            if not clean_str(row[0]) == MAGIC_DATA_STRING:
                if MAGIC_DATE_STRING in clean_str(row[0]):
                    date = row[1]

                    if isinstance(row[1], datetime.datetime):
                        date = row[1].date()
                if MAGIC_FREQ_STRING in clean_str(row[0]):
                    print("A")
                    freq = row[1]
                if MAGIC_DOSG_STRING in clean_str(row[0]):
                    print("B")
                    dosg = row[1]
                if MAGIC_EFFS_STRING in clean_str(row[0]):
                    if clean_str(row[1]) != "no":
                        effs = row[2]
                    else:
                        obvs = row[2]
                
                continue
            
            row = [parse_ranges(v) for v in row[1:]]
            desc_row = [clean_str(name) for name in sheet_values[i - 1][1:] if name]
            row_data = {desc_row[i]: row[i] for i in range(len(desc_row))}

            if row_data and [v for v in row if v]:
                row_data["fecha"] = date
                date = ""
                
                row_data["frecuencia"] = freq 
                freq = ""

                row_data["dosis"] = dosg
                dosg = ""

                row_data["efectos adversos"] = effs
                effs = ""

                row_data["observaciones"] = obvs
                obvs = ""

                data[patient_id].append(row_data)

        if trim_length != -1:
            if len(data[patient_id]) < trim_length:
                del data[patient_id]
            else:
                data[patient_id] = data[patient_id][:trim_length]

    for dataset in data.values():
        for query in dataset:
            for key in query.keys():
                keys.add(clean_str(key))
    
    output = "output.csv"
    if trim_length != -1:
        output = f"output-{trim_length}-consultas.csv"

    with open(output, "w", newline="") as f:
        print(f"Writing CSV output to {output}")

        writer = csv.writer(f, delimiter="\t")

        header = ["id", "paciente", "tiempo"]
        header.extend(keys)

        writer.writerow(header)
            
        i = 0
        for patient_id in data.keys():
            for (j, values) in enumerate(data[patient_id]):
                row = [i, patient_id, j]

                for key in keys:
                    if key in values and values[key] and not values[key] in ["s/d", "n/d"]:
                        row.append(values[key])
                    else:
                        row.append("ND")

                writer.writerow(row)
                i += 1


if __name__ == "__main__":
    if len(sys.argv) < 2 or len(sys.argv) > 3:
        print("ERROR: Incorrect usage")
        print("Usage: python3 convert.py [file_name] [required_queries (disabled by default)]")
        exit()
    
    filename = sys.argv[1]

    trim_length = -1
    if len(sys.argv) == 3:
        trim_length = int(sys.argv[2])

    main(filename, trim_length)
